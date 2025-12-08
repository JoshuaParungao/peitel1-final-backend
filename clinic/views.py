from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
import io
import csv
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from django.conf import settings
from django.contrib.staticfiles import finders
from reportlab.lib import colors
from reportlab.lib.units import mm
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth import login as auth_login, logout as auth_logout
from django.contrib.auth.models import User
from django.urls import reverse
from django.utils import timezone
from datetime import timedelta
from .models import Patient, Service, Invoice, InvoiceItem, StaffProfile
from .forms import PatientForm, ServiceForm, InvoiceForm

def superuser_required(view_func):
    return user_passes_test(lambda u: u.is_authenticated and u.is_superuser, login_url='login')(view_func)

# 1. Authentication
def login_view(request):
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect('clinic:dashboard')
    if request.method == 'POST':
        from django.contrib.auth import authenticate
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user and user.is_superuser:
            auth_login(request, user)
            return redirect('clinic:dashboard')
        return render(request, 'clinic/login.html', {'error': 'Invalid credentials or not superuser'})
    return render(request, 'clinic/login.html')

@superuser_required
def logout_view(request):
    auth_logout(request)
    return redirect('clinic:dashboard')

# 2. Dashboard
@superuser_required
def dashboard(request):
    context = {
        'total_patients': Patient.objects.filter(is_archived=False).count(),
        'total_services': Service.objects.filter(is_archived=False).count(),
        'total_invoices': Invoice.objects.filter(is_archived=False).count(),
        'total_staff': User.objects.filter(is_staff=True, is_active=True).count(),
    }
    return render(request, 'clinic/dashboard.html', context)

# 3. Patients Module
@superuser_required
def patients_list(request):
    patients = Patient.objects.filter(is_archived=False)
    return render(request, 'clinic/patient_list.html', {'patients': patients, 'back_url': reverse('clinic:dashboard')})

@superuser_required
def patient_detail(request, pk):
    patient = get_object_or_404(Patient, pk=pk)

    # Gather invoices and service usage for this patient
    invoices = Invoice.objects.filter(patient=patient).order_by('-date_created')

    # Summarize services from invoice items
    service_items = InvoiceItem.objects.filter(invoice__patient=patient)
    services_summary = {}
    for si in service_items:
        name = si.service_name_at_time
        qty = si.quantity or 0
        total = si.total_price() or 0
        if name in services_summary:
            services_summary[name]['quantity'] += qty
            services_summary[name]['total'] += float(total)
        else:
            services_summary[name] = {'name': name, 'quantity': qty, 'total': float(total)}

    services_summary_list = list(services_summary.values())

    return render(request, 'clinic/patient_detail.html', {
        'patient': patient,
        'invoices': invoices,
        'services_summary': services_summary_list,
        'back_url': reverse('clinic:dashboard')
    })

@superuser_required
def patient_create(request):
    if request.method == 'POST':
        form = PatientForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('clinic:patients_list')
    else:
        form = PatientForm()
    return render(request, 'clinic/patient_form.html', {'form': form, 'back_url': reverse('clinic:dashboard')})

@superuser_required
def patient_update(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            form.save()
            return redirect('clinic:patients_list')
    else:
        form = PatientForm(instance=patient)
    return render(request, 'clinic/patient_form.html', {'form': form, 'back_url': reverse('clinic:dashboard')})

@superuser_required
def patient_delete(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    # Soft-delete the patient and archive related invoices
    patient.is_archived = True
    patient.save()
    # Archive all invoices for this patient (idempotent)
    Invoice.objects.filter(patient=patient).update(is_archived=True)
    return redirect('clinic:patients_list')

# 4. Services Module
@superuser_required
def services_list(request):
    services = Service.objects.filter(is_archived=False)
    return render(request, 'clinic/service_list.html', {'services': services, 'back_url': reverse('clinic:dashboard')})

@superuser_required
def service_create(request):
    if request.method == 'POST':
        form = ServiceForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('clinic:services_list')
    else:
        form = ServiceForm()
    return render(request, 'clinic/service_form.html', {'form': form, 'back_url': reverse('clinic:dashboard')})

@superuser_required
def service_update(request, pk):
    service = get_object_or_404(Service, pk=pk)
    if request.method == 'POST':
        form = ServiceForm(request.POST, instance=service)
        if form.is_valid():
            form.save()
            return redirect('clinic:services_list')
    else:
        form = ServiceForm(instance=service)
    return render(request, 'clinic/service_form.html', {'form': form, 'back_url': reverse('clinic:dashboard')})

@superuser_required
def service_delete(request, pk):
    service = get_object_or_404(Service, pk=pk)
    # Soft-delete the service and archive any invoices that include this service
    service.is_archived = True
    service.save()
    # Archive invoices that reference this service via InvoiceItem
    Invoice.objects.filter(items__service=service).update(is_archived=True)
    return redirect('clinic:services_list')

# 5. Invoices Module
@superuser_required
def invoices_list(request):
    invoices = Invoice.objects.filter(is_archived=False)
    return render(request, 'clinic/invoice_list.html', {'invoices': invoices})

@superuser_required
def invoice_detail(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    return render(request, 'clinic/invoice_detail.html', {'invoice': invoice})

@superuser_required
def invoice_create(request):
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        if form.is_valid():
            invoice = form.save(commit=False)
            invoice.created_by = request.user
            invoice.save()
            return redirect('clinic:invoice_detail', pk=invoice.pk)
    else:
        form = InvoiceForm()
    return render(request, 'clinic/invoice_form.html', {'form': form, 'back_url': reverse('clinic:dashboard')})

@superuser_required
def invoice_update(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    if request.method == 'POST':
        form = InvoiceForm(request.POST, instance=invoice)
        if form.is_valid():
            form.save()
            return redirect('clinic:invoice_detail', pk=invoice.pk)
    else:
        form = InvoiceForm(instance=invoice)
    return render(request, 'clinic/invoice_form.html', {'form': form, 'invoice': invoice, 'back_url': reverse('clinic:dashboard')})

@superuser_required
def invoice_delete(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    invoice.is_archived = True
    invoice.save()
    return redirect('clinic:invoices_list')


@superuser_required
def invoice_pdf(request, pk):
    """Generate a PDF for a single invoice (download)."""
    invoice = get_object_or_404(Invoice, pk=pk)
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    x = 40
    y = height - 40
    p.setFont('Helvetica-Bold', 16)
    p.drawString(x, y, f"Invoice #{invoice.pk}")
    p.setFont('Helvetica', 10)
    y -= 24
    p.drawString(x, y, f"Patient: {invoice.patient.first_name} {invoice.patient.last_name}")
    y -= 14
    p.drawString(x, y, f"Date: {invoice.date_created.strftime('%Y-%m-%d %H:%M')}")
    y -= 20

    # Table header
    p.setFont('Helvetica-Bold', 11)
    p.drawString(x, y, 'Item')
    p.drawString(x+260, y, 'Price')
    p.drawString(x+340, y, 'Qty')
    p.drawString(x+400, y, 'Total')
    y -= 14
    p.setFont('Helvetica', 10)

    total = 0
    for item in invoice.items.all():
        if y < 80:
            p.showPage()
            y = height - 40
        p.drawString(x, y, item.service_name_at_time)
        p.drawRightString(x+320, y, f"₱{float(item.price_at_time):,.2f}")
        p.drawRightString(x+390, y, str(item.quantity))
        line_total = float(item.price_at_time) * item.quantity
        p.drawRightString(x+480, y, f"₱{line_total:,.2f}")
        total += line_total
        y -= 14

    y -= 8
    p.setFont('Helvetica-Bold', 12)
    p.drawString(x, y, f"Total: ₱{total:,.2f}")

    p.showPage()
    p.save()
    buffer.seek(0)
    resp = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    # Use inline so browsers open the PDF directly allowing printing (instead of forcing download)
    resp['Content-Disposition'] = f'inline; filename="invoice_{invoice.pk}.pdf"'
    return resp


@superuser_required
def sales_summary_csv(request):
    """Export invoices (not archived) as CSV with optional date-range filter and a summary row.
    Query params: ?start=YYYY-MM-DD&end=YYYY-MM-DD
    """
    invoices = Invoice.objects.filter(is_archived=False).order_by('date_created')
    # date range filter
    start = request.GET.get('start')
    end = request.GET.get('end')
    from datetime import date
    try:
        if start:
            start_date = date.fromisoformat(start)
            invoices = invoices.filter(date_created__date__gte=start_date)
        if end:
            end_date = date.fromisoformat(end)
            invoices = invoices.filter(date_created__date__lte=end_date)
    except Exception:
        # ignore parse errors and return full set
        pass

    total_sales = 0
    total_count = invoices.count()

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    # header
    writer.writerow(['Invoice ID', 'Patient', 'Date', 'Is Paid', 'Created By', 'Total Amount'])
    for inv in invoices:
        amount = sum(float(it.price_at_time) * it.quantity for it in inv.items.all())
        total_sales += amount
        writer.writerow([inv.pk, f"{inv.patient.first_name} {inv.patient.last_name}", inv.date_created.strftime('%Y-%m-%d %H:%M'), 'Yes' if inv.is_paid else 'No', (inv.created_by.get_full_name() if inv.created_by else ''), f"{amount:.2f}"])

    # summary row
    writer.writerow([])
    writer.writerow(['TOTAL_INVOICES', total_count])
    writer.writerow(['TOTAL_SALES', f"{total_sales:.2f}"])

    resp = HttpResponse(buffer.getvalue(), content_type='text/csv')
    filename = 'sales_summary'
    if start or end:
        filename += f"_{start or ''}_{end or ''}"
    resp['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    return resp


@superuser_required
def sales_summary_pdf(request):
    """Generate a PDF summarizing sales (list + totals) with optional date-range filter.
    Query params: ?start=YYYY-MM-DD&end=YYYY-MM-DD
    """
    invoices = Invoice.objects.filter(is_archived=False).order_by('date_created')
    # date range filter
    start = request.GET.get('start')
    end = request.GET.get('end')
    from datetime import date
    try:
        if start:
            start_date = date.fromisoformat(start)
            invoices = invoices.filter(date_created__date__gte=start_date)
        if end:
            end_date = date.fromisoformat(end)
            invoices = invoices.filter(date_created__date__lte=end_date)
    except Exception:
        pass

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    margin_x = 20 * mm
    margin_y = 20 * mm
    usable_width = width - 2 * margin_x
    x = margin_x
    y = height - margin_y

    # Header: optionally draw logo
    clinic_name = getattr(settings, 'CLINIC_NAME', 'Dental Clinic')
    clinic_addr = getattr(settings, 'CLINIC_ADDRESS', '')
    logo_path = None
    try:
        logo_path = finders.find('clinic/img/logo.png') or finders.find('clinic/logo.png')
    except Exception:
        logo_path = None
    if logo_path:
        try:
            img_w = 30 * mm
            img_h = 30 * mm
            p.drawImage(logo_path, x, y - img_h, width=img_w, height=img_h, preserveAspectRatio=True, mask='auto')
        except Exception:
            logo_path = None

    # Clinic header text
    header_x = x + (35 * mm if logo_path else 0)
    p.setFont('Helvetica-Bold', 14)
    p.drawString(header_x, y - 6, clinic_name)
    p.setFont('Helvetica', 9)
    p.drawString(header_x, y - 22, clinic_addr)
    y -= (40 if logo_path else 24)

    # Title and date range info
    p.setFont('Helvetica-Bold', 12)
    p.drawString(x, y, 'Sales Summary')
    if start or end:
        range_text = 'Range:'
        if start:
            range_text += f' {start}'
        if end:
            range_text += f' to {end}'
        p.setFont('Helvetica', 9)
        p.drawString(x + 120 * mm, y, range_text)
    y -= 12

    # Table header
    table_x = x
    col_invoice = table_x
    col_patient = table_x + 30 * mm
    col_date = table_x + 120 * mm
    col_total = table_x + usable_width - 30 * mm

    p.setFont('Helvetica-Bold', 10)
    p.drawString(col_invoice, y, 'Invoice')
    p.drawString(col_patient, y, 'Patient')
    p.drawString(col_date, y, 'Date')
    p.drawRightString(col_total + 30 * mm, y, 'Total')
    y -= 8

    # draw line under header
    p.setStrokeColor(colors.grey)
    p.setLineWidth(0.5)
    p.line(table_x, y, table_x + usable_width, y)
    y -= 8

    p.setFont('Helvetica', 9)
    total_sales = 0
    for inv in invoices:
        if y < margin_y + 30:
            p.showPage()
            y = height - margin_y
            # redraw header on new page
            p.setFont('Helvetica-Bold', 12)
            p.drawString(x, y, 'Sales Summary (continued)')
            y -= 14
            p.setFont('Helvetica-Bold', 10)
            p.drawString(col_invoice, y, 'Invoice')
            p.drawString(col_patient, y, 'Patient')
            p.drawString(col_date, y, 'Date')
            p.drawRightString(col_total + 30 * mm, y, 'Total')
            y -= 12

        amount = sum(float(it.price_at_time) * it.quantity for it in inv.items.all())
        total_sales += amount
        # draw row
        p.setFillColor(colors.black)
        p.drawString(col_invoice, y, str(inv.pk))
        patient_name = f"{inv.patient.first_name} {inv.patient.last_name}"
        p.drawString(col_patient, y, patient_name[:40])
        p.drawString(col_date, y, inv.date_created.strftime('%Y-%m-%d'))
        p.drawRightString(col_total + 30 * mm, y, f"₱{amount:,.2f}")
        y -= 12

    # footer totals
    y -= 8
    p.setStrokeColor(colors.grey)
    p.line(table_x, y, table_x + usable_width, y)
    y -= 12
    p.setFont('Helvetica-Bold', 11)
    p.drawString(table_x, y, f"Total Invoices: {invoices.count()}")
    p.drawRightString(table_x + usable_width, y, f"Total Sales: ₱{total_sales:,.2f}")

    p.showPage()
    p.save()
    buffer.seek(0)
    resp = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = 'sales_summary'
    if start or end:
        filename += f"_{start or ''}_{end or ''}"
    # Open inline so browsers display the PDF for printing
    resp['Content-Disposition'] = f'inline; filename="{filename}.pdf"'
    return resp


@superuser_required
def sales_summary_xlsx(request):
    """Export all invoices (not archived) as an Excel .xlsx file."""
    if openpyxl is None:
        return HttpResponse('openpyxl is not installed. Install with `pip install openpyxl`', status=500)
    invoices = Invoice.objects.filter(is_archived=False).order_by('date_created')
    # date range filter
    start = request.GET.get('start')
    end = request.GET.get('end')
    from datetime import date
    try:
        if start:
            start_date = date.fromisoformat(start)
            invoices = invoices.filter(date_created__date__gte=start_date)
        if end:
            end_date = date.fromisoformat(end)
            invoices = invoices.filter(date_created__date__lte=end_date)
    except Exception:
        pass
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales'

    headers = ['Invoice ID', 'Patient', 'Date', 'Is Paid', 'Created By', 'Total Amount']
    ws.append(headers)

    total_sales = 0
    for inv in invoices:
        amount = sum(float(it.price_at_time) * it.quantity for it in inv.items.all())
        total_sales += amount
        ws.append([inv.pk, f"{inv.patient.first_name} {inv.patient.last_name}", inv.date_created.strftime('%Y-%m-%d %H:%M'), 'Yes' if inv.is_paid else 'No', (inv.created_by.get_full_name() if inv.created_by else ''), float(f"{amount:.2f}")])

    # summary rows
    ws.append([])
    ws.append(['TOTAL_INVOICES', invoices.count()])
    ws.append(['TOTAL_SALES', float(f"{total_sales:.2f}")])

    # auto-size columns
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(i)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column].width = (max_length + 2)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    resp = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = 'sales_summary'
    if start or end:
        filename += f"_{start or ''}_{end or ''}"
    resp['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return resp

# 6. Archive Module
@superuser_required
def archive(request):
    archived_patients = Patient.objects.filter(is_archived=True)
    archived_services = Service.objects.filter(is_archived=True)
    archived_invoices = Invoice.objects.filter(is_archived=True)
    archived_staff = StaffProfile.objects.filter(is_archived=True)
    return render(request, 'clinic/archive.html', {
        'archived_patients': archived_patients,
        'archived_services': archived_services,
        'archived_invoices': archived_invoices,
        'archived_staff': archived_staff,
        'back_url': reverse('clinic:dashboard')
    })

# Restore functions
@superuser_required
def restore_patient(request, pk):
    patient = get_object_or_404(Patient, pk=pk, is_archived=True)
    patient.is_archived = False
    patient.save()
    return redirect('clinic:archive')

@superuser_required
def restore_service(request, pk):
    service = get_object_or_404(Service, pk=pk, is_archived=True)
    service.is_archived = False
    service.save()
    return redirect('clinic:archive')

@superuser_required
def restore_invoice(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk, is_archived=True)
    invoice.is_archived = False
    invoice.save()
    return redirect('clinic:archive')

# Permanent delete functions
@superuser_required
def delete_patient_permanent(request, pk):
    patient = get_object_or_404(Patient, pk=pk, is_archived=True)
    patient.delete()
    return redirect('clinic:archive')

@superuser_required
def delete_service_permanent(request, pk):
    service = get_object_or_404(Service, pk=pk, is_archived=True)
    service.delete()
    return redirect('clinic:archive')

@superuser_required
def delete_invoice_permanent(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk, is_archived=True)
    invoice.delete()
    return redirect('clinic:archive')

# 7. Staff Approval Module
@superuser_required
def staff_approval(request):
    pending_staff = User.objects.filter(is_staff=True, is_active=False)
    # If a user just registered, read session flag and surface a confirmation banner
    just_registered = None
    try:
        just_registered = request.session.pop('registered_username', None)
    except Exception:
        just_registered = None

    # If the username from session exists and is still pending, pass it to template
    just_registered_present = False
    if just_registered:
        just_registered_present = pending_staff.filter(username=just_registered).exists()

    context = {
        'pending_staff': pending_staff,
        'back_url': reverse('clinic:dashboard'),
        'just_registered': just_registered if just_registered_present else None,
    }
    return render(request, 'clinic/staff_approval.html', context)

@superuser_required
def approve_staff(request, pk):
    staff = get_object_or_404(User, pk=pk, is_staff=True)
    staff.is_active = True
    staff.save()
    # mark StaffProfile as approved when approving via admin
    try:
        profile = staff.staff_profile
        profile.approved = True
        profile.save()
    except Exception:
        pass
    return redirect('clinic:staff_approval')

@superuser_required
def reject_staff(request, pk):
    staff = get_object_or_404(User, pk=pk, is_staff=True)
    staff.delete()
    return redirect('clinic:staff_approval')


@superuser_required
def staff_delete(request, pk):
    staff = get_object_or_404(User, pk=pk, is_staff=True)
    # soft-archive: mark StaffProfile.is_archived and deactivate user
    profile = getattr(staff, 'staff_profile', None)
    if profile:
        profile.is_archived = True
        profile.save()
    staff.is_active = False
    staff.save()
    return redirect('clinic:staff_list')


@superuser_required
def restore_staff(request, pk):
    # pk is StaffProfile id
    profile = get_object_or_404(StaffProfile, pk=pk, is_archived=True)
    profile.is_archived = False
    profile.save()
    # reactivate user
    user = profile.user
    user.is_active = True
    user.save()
    return redirect('clinic:archive')


@superuser_required
def delete_staff_permanent(request, pk):
    # pk is StaffProfile id
    profile = get_object_or_404(StaffProfile, pk=pk, is_archived=True)
    user = profile.user
    profile.delete()
    if user:
        user.delete()
    return redirect('clinic:archive')

# 8. Staff List with Activity Tracking
@superuser_required
def staff_list(request):
    """Display all active staff (excluding superusers) with their activity and sales metrics"""
    # Filter only staff users, excluding superusers
    staff_list = User.objects.filter(is_staff=True, is_active=True, is_superuser=False)
    
    staff_data = []
    total_invoices_count = 0
    total_sales_amount = 0
    
    for staff in staff_list:
        total_invoices = Invoice.objects.filter(created_by=staff, is_archived=False).count()
        total_sales = sum(inv.total_amount() for inv in Invoice.objects.filter(created_by=staff, is_archived=False))
        last_activity = Invoice.objects.filter(created_by=staff).values_list('date_created', flat=True).order_by('-date_created').first()
        
        total_invoices_count += total_invoices
        total_sales_amount += total_sales
        
        staff_data.append({
            'user': staff,
            'total_invoices': total_invoices,
            'total_sales': total_sales or 0,
            'last_activity': last_activity,
            'profile': getattr(staff, 'staff_profile', None)
        })
    
    return render(request, 'clinic/staff_list.html', {
        'staff_data': staff_data,
        'total_invoices_count': total_invoices_count,
        'total_sales_amount': total_sales_amount,
        'back_url': reverse('clinic:dashboard')
    })


# --- Staff POS (mobile) -------------------------------------------------
def staff_required(view_func):
    return user_passes_test(lambda u: u.is_authenticated and u.is_staff and u.is_active and getattr(getattr(u, 'staff_profile', None), 'approved', False), login_url='clinic:staff_login')(view_func)


@staff_required
def staff_pos(request):
    """Simple mobile-friendly POS for approved staff: add/select patient, pick services, generate invoice."""
    services = Service.objects.filter(active=True, is_archived=False)
    if request.method == 'POST':
        # Determine patient: existing or new
        patient_id = request.POST.get('patient_id')
        if patient_id:
            patient = get_object_or_404(Patient, pk=patient_id)
        else:
            # Create new patient from minimal fields
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            contact_number = request.POST.get('contact_number')
            email = request.POST.get('email')
            address = request.POST.get('address')
            patient = Patient.objects.create(
                first_name=first_name or 'Unknown',
                last_name=last_name or 'Patient',
                contact_number=contact_number or '',
                email=email or '',
                address=address or '',
                created_by=request.user
            )

        # Create invoice
        invoice = Invoice.objects.create(patient=patient, created_by=request.user)

        # Collect service quantities from POST. Inputs are named service_<id>
        for svc in services:
            qty_raw = request.POST.get(f'service_{svc.id}')
            try:
                qty = int(qty_raw) if qty_raw else 0
            except ValueError:
                qty = 0
            if qty and qty > 0:
                InvoiceItem.objects.create(invoice=invoice, service=svc, quantity=qty)

        return redirect('clinic:invoice_detail', pk=invoice.pk)

    # GET: show POS interface
    return render(request, 'clinic/staff_pos.html', {'services': services})

# 9. Sales Analytics Module
@superuser_required
def sales_analytics(request):
    """Display sales analytics: daily, weekly, monthly, yearly"""
    today = timezone.now().date()
    
    # Daily sales
    daily_start = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.min.time()))
    daily_invoices = Invoice.objects.filter(date_created__gte=daily_start, is_archived=False)
    daily_sales = sum(inv.total_amount() for inv in daily_invoices)
    daily_count = daily_invoices.count()
    
    # Weekly sales (last 7 days)
    weekly_start = today - timedelta(days=7)
    weekly_start = timezone.make_aware(timezone.datetime.combine(weekly_start, timezone.datetime.min.time()))
    weekly_invoices = Invoice.objects.filter(date_created__gte=weekly_start, is_archived=False)
    weekly_sales = sum(inv.total_amount() for inv in weekly_invoices)
    weekly_count = weekly_invoices.count()
    
    # Monthly sales (current month)
    monthly_start = today.replace(day=1)
    monthly_start = timezone.make_aware(timezone.datetime.combine(monthly_start, timezone.datetime.min.time()))
    monthly_invoices = Invoice.objects.filter(date_created__gte=monthly_start, is_archived=False)
    monthly_sales = sum(inv.total_amount() for inv in monthly_invoices)
    monthly_count = monthly_invoices.count()
    
    # Yearly sales (current year)
    yearly_start = today.replace(month=1, day=1)
    yearly_start = timezone.make_aware(timezone.datetime.combine(yearly_start, timezone.datetime.min.time()))
    yearly_invoices = Invoice.objects.filter(date_created__gte=yearly_start, is_archived=False)
    yearly_sales = sum(inv.total_amount() for inv in yearly_invoices)
    yearly_count = yearly_invoices.count()
    
    # Staff-wise sales breakdown
    staff_sales = []
    for staff in User.objects.filter(is_staff=True, is_active=True):
        staff_invoices = Invoice.objects.filter(created_by=staff, is_archived=False)
        yearly_staff_sales = sum(inv.total_amount() for inv in staff_invoices.filter(date_created__gte=yearly_start))
        
        # Calculate percentage
        if yearly_sales > 0:
            percentage = (yearly_staff_sales / yearly_sales) * 100
        else:
            percentage = 0
        
        staff_sales.append({
            'staff': staff,
            'total_sales': sum(inv.total_amount() for inv in staff_invoices),
            'total_invoices': staff_invoices.count(),
            'daily': sum(inv.total_amount() for inv in staff_invoices.filter(date_created__gte=daily_start)),
            'weekly': sum(inv.total_amount() for inv in staff_invoices.filter(date_created__gte=weekly_start)),
            'monthly': sum(inv.total_amount() for inv in staff_invoices.filter(date_created__gte=monthly_start)),
            'yearly': yearly_staff_sales,
            'yearly_percentage': percentage,
        })
    
    context = {
        'daily_sales': daily_sales,
        'daily_count': daily_count,
        'weekly_sales': weekly_sales,
        'weekly_count': weekly_count,
        'monthly_sales': monthly_sales,
        'monthly_count': monthly_count,
        'yearly_sales': yearly_sales,
        'yearly_count': yearly_count,
        'staff_sales': staff_sales,
        'back_url': reverse('clinic:dashboard')
    }
    return render(request, 'clinic/sales_analytics.html', context)

# 10. Navigation / Back Button is handled via 'back_url' context in templates

# 11. Logout is handled above
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .decorators import frontend_login_required
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import User, Group
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from .models import Service, Patient, Invoice, InvoiceItem
from .forms import StaffRegistrationForm, PatientForm
import json

# DRF imports for API
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.response import Response
from rest_framework import status, permissions
from rest_framework.authentication import SessionAuthentication, TokenAuthentication
from rest_framework.authtoken.models import Token
from .serializers import ServiceSerializer, PatientSerializer, InvoiceSerializer
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


def landing_view(request):
    return render(request, 'clinic/landing.html')


def superuser_required(view_func):
    return user_passes_test(lambda u: u.is_authenticated and u.is_superuser, login_url='clinic:login')(view_func)


@superuser_required
def super_admin_dashboard(request):
    """Dashboard visible to superusers only."""
    total_patients = Patient.objects.count()
    total_invoices = Invoice.objects.filter(is_archived=False).count()
    total_staff = User.objects.filter(is_staff=True).count()
    recent_invoices = Invoice.objects.filter(is_archived=False).order_by('-date_created')[:10]
    return render(request, 'clinic/super_admin_dashboard.html', {
        'total_patients': total_patients,
        'total_invoices': total_invoices,
        'total_staff': total_staff,
        'recent_invoices': recent_invoices,
    })


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def api_health_check(request):
    """API health check - returns JSON"""
    return Response({
        'status': 'ok',
        'message': 'Dental Clinic POS API is running',
        'version': '1.0.0'
    }, status=status.HTTP_200_OK)


@superuser_required
def sales_view(request):
    """Superuser-only sales dashboard with summary cards and graphs."""
    # Basic KPIs
    from django.db.models import Sum
    from django.utils import timezone
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = (today_start - timezone.timedelta(days=now.weekday()))
    month_start = today_start.replace(day=1)
    year_start = today_start.replace(month=1, day=1)

    daily_total = Invoice.objects.filter(is_archived=False, date_created__gte=today_start).aggregate(total=Sum('items__price_at_time'))['total'] or 0
    weekly_total = Invoice.objects.filter(is_archived=False, date_created__gte=week_start).aggregate(total=Sum('items__price_at_time'))['total'] or 0
    monthly_total = Invoice.objects.filter(is_archived=False, date_created__gte=month_start).aggregate(total=Sum('items__price_at_time'))['total'] or 0
    yearly_total = Invoice.objects.filter(is_archived=False, date_created__gte=year_start).aggregate(total=Sum('items__price_at_time'))['total'] or 0

    return render(request, 'clinic/sales_admin.html', {
        'daily_total': float(daily_total),
        'weekly_total': float(weekly_total),
        'monthly_total': float(monthly_total),
        'yearly_total': float(yearly_total),
    })


def staff_register(request):
    try:
        if request.method == 'POST':
            form = StaffRegistrationForm(request.POST)
            if form.is_valid():
                user = form.save(commit=False)
                user.set_password(form.cleaned_data['password'])
                user.is_active = False  # Pending approval
                user.is_staff = True
                user.save()
                group, created = Group.objects.get_or_create(name='Frontdesk')
                user.groups.add(group)
                # Create StaffProfile for this user
                from .models import StaffProfile
                StaffProfile.objects.create(user=user, position='other')
                return render(request, 'clinic/registration_pending.html')
        else:
            form = StaffRegistrationForm()
        return render(request, 'clinic/staff_register.html', {'form': form})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return render(request, 'clinic/staff_register.html', {'form': StaffRegistrationForm(), 'error': str(e)})


@superuser_required
def archived_staff_list(request):
    archived = User.objects.filter(is_staff=True, staff_profile__is_archived=True)
    return render(request, 'clinic/archived_staff_list.html', {'archived': archived})


@superuser_required
def new_registrations(request):
    """List newly registered staff awaiting approval."""
    pending = User.objects.filter(is_staff=True, is_active=False).select_related('staff_profile')
    return render(request, 'clinic/new_registrations.html', {'pending': pending})


@superuser_required
def invoice_list(request):
    invoices = Invoice.objects.filter(is_archived=False).order_by('-date_created')
    return render(request, 'clinic/invoice_list.html', {'invoices': invoices})


@superuser_required
def edit_invoice(request, invoice_id):
    try:
        invoice = Invoice.objects.get(id=invoice_id)
    except Invoice.DoesNotExist:
        return redirect('clinic:invoice_list')

    if request.method == 'POST':
        # Allow toggling paid status and simple note updates
        is_paid = request.POST.get('is_paid') == 'on'
        invoice.is_paid = is_paid
        invoice.save()
        return redirect('clinic:invoice_list')

    return render(request, 'clinic/edit_invoice.html', {'invoice': invoice})


@superuser_required
def archived_invoices_list(request):
    invoices = Invoice.objects.filter(is_archived=True).order_by('-date_created')
    return render(request, 'clinic/archived_invoices_list.html', {'invoices': invoices})


@superuser_required
def archive_invoice(request, invoice_id):
    try:
        inv = Invoice.objects.get(id=invoice_id)
        inv.is_archived = True
        inv.save()
    except Invoice.DoesNotExist:
        pass
    return redirect('clinic:invoice_list')


@superuser_required
def archive_staff(request, user_id):
    try:
        user = User.objects.get(id=user_id)
        try:
            sp = user.staff_profile
            sp.is_archived = True
            sp.save()
        except Exception:
            pass
        # Optionally deactivate account
        user.is_active = False
        user.save()
    except User.DoesNotExist:
        pass
    return redirect('clinic:staff_list')


@frontend_login_required
def select_services(request, patient_id):
    """Select services for a new patient and create invoice"""
    patient = Patient.objects.get(id=patient_id)
    services = Service.objects.filter(active=True)

    # Ensure Medical Certificate service exists in DB so it's visible in frontend
    try:
        mc_defaults = {
            'name': 'Medical Certificate',
            'price': Service.DEFAULT_PRICES.get('MEDICAL_CERTIFICATE', 300),
            'active': True,
        }
        Service.objects.get_or_create(category='MEDICAL_CERTIFICATE', defaults=mc_defaults)
        # refresh queryset
        services = Service.objects.filter(active=True)
    except Exception:
        # If something goes wrong creating the record, continue without failing the view
        pass
    
    if request.method == 'POST':
        service_ids = request.POST.getlist('service_id')
        quantities = request.POST.getlist('quantity')
        
        if service_ids:
            invoice = Invoice.objects.create(patient=patient, created_by=request.user if request.user.is_authenticated else None)
            for s_id, qty in zip(service_ids, quantities):
                service = Service.objects.get(id=s_id)
                InvoiceItem.objects.create(
                    invoice=invoice,
                    service=service,
                    quantity=int(qty) if qty else 1,
                    price_at_time=service.price,
                    service_name_at_time=service.name
                )
            return redirect('clinic:invoice_detail', pk=invoice.id)
    
    return render(request, 'clinic/select_services.html', {'patient': patient, 'services': services})


@frontend_login_required
def service_list(request):
    services = Service.objects.filter(active=True)
    return render(request, 'clinic/service_list.html', {'services': services})


@superuser_required
def add_service(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        category = request.POST.get('category')
        price = request.POST.get('price')
        description = request.POST.get('description', '')
        Service.objects.create(name=name, category=category, price=price, description=description, active=True)
        return redirect('clinic:service_list')
    return render(request, 'clinic/add_service.html', {'categories': Service.DENTAL_CATEGORIES})


@superuser_required
def edit_service(request, service_id):
    try:
        service = Service.objects.get(id=service_id)
    except Service.DoesNotExist:
        return redirect('clinic:service_list')
    if request.method == 'POST':
        service.name = request.POST.get('name')
        service.category = request.POST.get('category')
        service.price = request.POST.get('price')
        service.description = request.POST.get('description', '')
        service.save()
        return redirect('clinic:service_list')
    return render(request, 'clinic/edit_service.html', {'service': service, 'categories': Service.DENTAL_CATEGORIES})


@superuser_required
def delete_service(request, service_id):
    try:
        service = Service.objects.get(id=service_id)
        # Soft-delete by marking inactive
        service.active = False
        service.save()
    except Service.DoesNotExist:
        pass
    return redirect('clinic:service_list')


@superuser_required
def patient_list(request):
    patients = Patient.objects.all().order_by('-created_at')
    return render(request, 'clinic/patient_list.html', {'patients': patients})


@frontend_login_required
def add_patient(request):
    """Add new patient to system"""
    if request.method == 'POST':
        form = PatientForm(request.POST)
        if form.is_valid():
            patient = form.save(commit=False)
            # Record which user created this patient
            if request.user.is_authenticated:
                patient.created_by = request.user
            patient.save()
            return redirect('clinic:select_services', patient_id=patient.id)
    else:
        form = PatientForm()
    return render(request, 'clinic/add_patient.html', {'form': form})


@frontend_login_required
def create_invoice(request):
    patients = Patient.objects.all()
    services = Service.objects.filter(active=True)
    if request.method == 'POST':
        patient_id = request.POST.get('patient')
        service_ids = request.POST.getlist('service_id')
        quantities = request.POST.getlist('quantity')
        patient = Patient.objects.get(id=patient_id)
        invoice = Invoice.objects.create(patient=patient, created_by=request.user if request.user.is_authenticated else None)
        for s_id, qty in zip(service_ids, quantities):
            service = Service.objects.get(id=s_id)
            InvoiceItem.objects.create(
                invoice=invoice,
                service=service,
                quantity=int(qty),
                price_at_time=service.price,
                service_name_at_time=service.name
            )
        return redirect('clinic:invoice_detail', pk=invoice.id)
    return render(request, 'clinic/create_invoice.html', {'patients': patients, 'services': services})


@frontend_login_required
def invoice_detail(request, pk):
    invoice = Invoice.objects.get(id=pk)
    # Show invoice detail UI
    return render(request, 'clinic/invoice_detail.html', {'invoice': invoice})


@frontend_login_required
def download_invoice_pdf(request, pk):
    """Download invoice as PDF"""
    invoice = Invoice.objects.get(id=pk)
    
    if not HAS_REPORTLAB:
        # Fallback: return HTML that browser can print to PDF
        html_string = render_to_string('clinic/invoice_pdf.html', {'invoice': invoice})
        return HttpResponse(html_string, content_type='text/html')
    
    # Generate PDF using ReportLab
    from io import BytesIO
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#198754'),
        spaceAfter=10,
        alignment=1
    )
    elements.append(Paragraph('🦷 INVOICE', title_style))
    elements.append(Paragraph(f'Invoice #{invoice.id}', ParagraphStyle('subtitle', parent=styles['Normal'], fontSize=12, alignment=1)))
    elements.append(Spacer(1, 0.3*inch))
    
    # Patient Info
    patient_style = ParagraphStyle('patient', parent=styles['Normal'], fontSize=10, leading=12)
    elements.append(Paragraph(f'<b>Patient:</b> {invoice.patient.first_name} {invoice.patient.last_name}', patient_style))
    elements.append(Paragraph(f'<b>Contact:</b> {invoice.patient.contact_number}', patient_style))
    elements.append(Paragraph(f'<b>Date:</b> {invoice.date_created.strftime("%B %d, %Y")}', patient_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Services Table
    table_data = [['Service', 'Price', 'Qty', 'Amount']]
    for item in invoice.items.all():
        table_data.append([
            item.service_name_at_time,
            f'₱{float(item.price_at_time):.2f}',
            str(item.quantity),
            f'₱{float(item.total_price()):.2f}'
        ])
    
    # Add total row
    total_amount = sum(float(item.total_price()) for item in invoice.items.all())
    table_data.append(['', '', 'TOTAL:', f'₱{total_amount:.2f}'])
    
    table = Table(table_data, colWidths=[2.5*inch, 1*inch, 0.8*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 1), (-1, -2), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -2), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Footer
    footer_style = ParagraphStyle('footer', parent=styles['Normal'], fontSize=9, textColor=colors.grey, alignment=1)
    elements.append(Paragraph('Thank you for choosing our dental clinic!', footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Invoice_{invoice.id}_{invoice.patient.first_name}.pdf"'
    
    return response


# ----------------------
# DRF JSON API endpoints for React Native + axios
# ----------------------


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def api_services_list(request):
    services = Service.objects.filter(active=True)
    serializer = ServiceSerializer(services, many=True)
    return Response({'services': serializer.data})


@api_view(['POST'])
@authentication_classes([TokenAuthentication, SessionAuthentication])
@permission_classes([permissions.IsAuthenticated])
def api_add_patient(request):
    serializer = PatientSerializer(data=request.data)
    if serializer.is_valid():
        patient = serializer.save(created_by=request.user if request.user and request.user.is_authenticated else None)
        return Response({'id': patient.id, 'first_name': patient.first_name, 'last_name': patient.last_name}, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@authentication_classes([TokenAuthentication, SessionAuthentication])
@permission_classes([permissions.IsAuthenticated])
def api_create_invoice(request):
    """Expected JSON: {"patient_id": int, "items": [{"service_id": int, "quantity": int}, ...]}"""
    patient_id = request.data.get('patient_id')
    items = request.data.get('items', [])
    if not patient_id or not isinstance(items, list) or len(items) == 0:
        return Response({'error': 'patient_id and items are required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        patient = Patient.objects.get(id=patient_id)
    except Patient.DoesNotExist:
        return Response({'error': 'Patient not found'}, status=status.HTTP_404_NOT_FOUND)

    invoice = Invoice.objects.create(patient=patient, created_by=request.user if request.user and request.user.is_authenticated else None)
    created_items = []
    for it in items:
        service_id = it.get('service_id')
        qty = int(it.get('quantity') or 1)
        try:
            service = Service.objects.get(id=service_id)
        except Service.DoesNotExist:
            invoice.delete()
            return Response({'error': f'Service {service_id} not found'}, status=status.HTTP_404_NOT_FOUND)

        ii = InvoiceItem.objects.create(
            invoice=invoice,
            service=service,
            quantity=qty,
            price_at_time=service.price,
            service_name_at_time=service.name,
        )
        created_items.append({'service_id': service.id, 'quantity': ii.quantity, 'price': float(ii.price_at_time)})

    total = float(invoice.total_amount())
    return Response({'invoice_id': invoice.id, 'total': total, 'items': created_items}, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def api_invoice_detail(request, pk):
    try:
        invoice = Invoice.objects.get(id=pk)
    except Invoice.DoesNotExist:
        return Response({'error': 'Invoice not found'}, status=status.HTTP_404_NOT_FOUND)

    serializer = InvoiceSerializer(invoice)
    # Add computed total and format items
    invoice_data = serializer.data
    items = []
    for it in invoice.items.all():
        items.append({
            'id': it.id,
            'service_id': it.service.id,
            'service_name': it.service_name_at_time,
            'price': float(it.price_at_time),
            'quantity': it.quantity,
            'amount': float(it.total_price()),
        })
    invoice_data['items'] = items
    invoice_data['total'] = float(invoice.total_amount())
    invoice_data['patient'] = {
        'id': invoice.patient.id,
        'first_name': invoice.patient.first_name,
        'last_name': invoice.patient.last_name,
        'contact_number': invoice.patient.contact_number,
    }
    return Response({'invoice': invoice_data})


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def api_login(request):
    try:
        # Handle both JSON and form data
        if hasattr(request, 'data'):
            data = request.data
        else:
            data = request.POST
        
        username = data.get('username') if hasattr(data, 'get') else getattr(data, 'username', None)
        password = data.get('password') if hasattr(data, 'get') else getattr(data, 'password', None)
        
        if not username or not password:
            return Response({'error': 'username and password required'}, status=status.HTTP_400_BAD_REQUEST)

        user = authenticate(request, username=username, password=password)
        if user is None:
            return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)
        if not user.is_active:
            return Response({'error': 'Account inactive'}, status=status.HTTP_403_FORBIDDEN)

        token, created = Token.objects.get_or_create(user=user)
        
        # Get position if staff profile exists
        position = None
        try:
            if hasattr(user, 'staff_profile') and user.staff_profile:
                position = user.staff_profile.get_position_display()
        except Exception as pos_e:
            # If staff profile doesn't exist, just set position to None
            position = 'admin' if user.is_superuser else None

        return Response({
            'token': token.key,
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'position': position
        })
    except Exception as e:
        import traceback
        error_msg = str(e)
        error_traceback = traceback.format_exc()
        print(f"LOGIN ERROR: {error_msg}")
        print(error_traceback)
        return Response({
            'error': 'Login failed',
            'detail': error_msg
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def api_register(request):
    """Register a new staff user (pending approval)"""
    try:
        # Handle both JSON and form data
        if hasattr(request, 'data'):
            data = request.data
        else:
            data = request.POST
        
        username = data.get('username') if hasattr(data, 'get') else getattr(data, 'username', None)
        password = data.get('password') if hasattr(data, 'get') else getattr(data, 'password', None)
        email = data.get('email') if hasattr(data, 'get') else getattr(data, 'email', None)
        first_name = data.get('first_name', '') if hasattr(data, 'get') else getattr(data, 'first_name', '')
        last_name = data.get('last_name', '') if hasattr(data, 'get') else getattr(data, 'last_name', '')
        position = data.get('position', 'other') if hasattr(data, 'get') else getattr(data, 'position', 'other')

        if not username or not password or not email:
            return Response({'error': 'username, password, and email are required'}, status=status.HTTP_400_BAD_REQUEST)

        # Check if user already exists
        if User.objects.filter(username=username).exists():
            return Response({'error': 'Username already taken'}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(email=email).exists():
            return Response({'error': 'Email already registered'}, status=status.HTTP_400_BAD_REQUEST)

        # Create user (inactive, pending approval)
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            is_active=False,
            is_staff=True
        )

        # Add to Frontdesk group
        group, _ = Group.objects.get_or_create(name='Frontdesk')
        user.groups.add(group)

        # Create staff profile with position
        from .models import StaffProfile
        StaffProfile.objects.create(user=user, position=position)

        return Response({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'position': position,
            'message': 'Registration successful. Awaiting admin approval.'
        }, status=status.HTTP_201_CREATED)
    except Exception as e:
        import traceback
        error_msg = str(e)
        error_traceback = traceback.format_exc()
        print(f"REGISTER ERROR: {error_msg}")
        print(error_traceback)
        return Response({
            'error': 'Registration failed',
            'detail': error_msg
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@authentication_classes([TokenAuthentication, SessionAuthentication])
@permission_classes([permissions.IsAuthenticated])
def api_logout(request):
    Token.objects.filter(user=request.user).delete()
    try:
        auth_logout(request)
    except Exception:
        pass
    return Response({'ok': True})


@api_view(['GET'])
@authentication_classes([TokenAuthentication, SessionAuthentication])
@permission_classes([permissions.IsAuthenticated])
def api_staff_activity(request):
    """Retrieve all staff activity - patients added, invoices created, revenue"""
    from .models import StaffProfile
    from django.db.models import Sum
    
    staff_profiles = StaffProfile.objects.select_related('user').all()
    
    activity_data = []
    for profile in staff_profiles:
        # Get patients added by this staff member
        patients = Patient.objects.filter(created_by=profile.user).order_by('-created_at')
        
        # Get invoices for those patients
        patient_ids = patients.values_list('id', flat=True)
        invoices = Invoice.objects.filter(patient_id__in=patient_ids).order_by('-date_created')
        
        # Calculate totals
        total_revenue = invoices.aggregate(Sum('items__price_at_time'))['items__price_at_time__sum'] or 0
        paid_invoices = invoices.filter(is_paid=True).count()
        pending_invoices = invoices.filter(is_paid=False).count()
        
        # Recent patients
        recent_patients = [
            {
                'id': p.id,
                'first_name': p.first_name,
                'last_name': p.last_name,
                'email': p.email,
                'phone': p.phone,
                'added_date': p.created_at.isoformat(),
                'invoices_count': p.invoices.count(),
                'total_spent': sum(inv.total_amount() for inv in p.invoices.all())
            }
            for p in patients[:10]
        ]
        
        # Recent invoices
        recent_invoices = [
            {
                'id': inv.id,
                'patient_name': f"{inv.patient.first_name} {inv.patient.last_name}",
                'amount': inv.total_amount(),
                'date_created': inv.date_created.isoformat(),
                'is_paid': inv.is_paid,
                'items_count': inv.items.count()
            }
            for inv in invoices[:10]
        ]
        
        activity_data.append({
            'staff_id': profile.id,
            'username': profile.user.username,
            'full_name': f"{profile.user.first_name} {profile.user.last_name}".strip() or "(No name)",
            'position': profile.get_position_display(),
            'email': profile.user.email,
            'created_at': profile.created_at.isoformat(),
            'statistics': {
                'patients_added': patients.count(),
                'total_invoices': invoices.count(),
                'total_revenue': float(total_revenue),
                'paid_invoices': paid_invoices,
                'pending_invoices': pending_invoices
            },
            'recent_patients': recent_patients,
            'recent_invoices': recent_invoices
        })
    
    return Response({
        'count': len(activity_data),
        'staff': activity_data
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
@authentication_classes([TokenAuthentication, SessionAuthentication])
@permission_classes([permissions.IsAuthenticated])
def api_staff_detail(request, staff_id):
    """Retrieve detailed activity for a specific staff member"""
    from .models import StaffProfile
    from django.db.models import Sum
    
    try:
        profile = StaffProfile.objects.select_related('user').get(id=staff_id)
    except StaffProfile.DoesNotExist:
        return Response({'error': 'Staff not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Get patients added by this staff member
    patients = Patient.objects.filter(created_by=profile.user).order_by('-created_at')
    
    # Get invoices for those patients
    patient_ids = patients.values_list('id', flat=True)
    invoices = Invoice.objects.filter(patient_id__in=patient_ids).order_by('-date_created')
    
    # Calculate totals
    total_revenue = invoices.aggregate(Sum('items__price_at_time'))['items__price_at_time__sum'] or 0
    paid_invoices = invoices.filter(is_paid=True).count()
    pending_invoices = invoices.filter(is_paid=False).count()
    
    # All patients (with pagination option)
    all_patients = [
        {
            'id': p.id,
            'first_name': p.first_name,
            'last_name': p.last_name,
            'email': p.email,
            'phone': p.phone,
            'added_date': p.created_at.isoformat(),
            'invoices_count': p.invoices.count(),
            'total_spent': sum(inv.total_amount() for inv in p.invoices.all())
        }
        for p in patients
    ]
    
    # All invoices (with pagination option)
    all_invoices = [
        {
            'id': inv.id,
            'patient_id': inv.patient_id,
            'patient_name': f"{inv.patient.first_name} {inv.patient.last_name}",
            'amount': inv.total_amount(),
            'date_created': inv.date_created.isoformat(),
            'is_paid': inv.is_paid,
            'items_count': inv.items.count(),
            'items': [
                {
                    'service': item.service.name,
                    'quantity': item.quantity,
                    'price': item.price_at_time
                }
                for item in inv.items.all()
            ]
        }
        for inv in invoices
    ]
    
    return Response({
        'staff_id': profile.id,
        'username': profile.user.username,
        'full_name': f"{profile.user.first_name} {profile.user.last_name}".strip() or "(No name)",
        'position': profile.get_position_display(),
        'email': profile.user.email,
        'created_at': profile.created_at.isoformat(),
        'statistics': {
            'total_patients': patients.count(),
            'total_invoices': invoices.count(),
            'total_revenue': float(total_revenue),
            'paid_invoices': paid_invoices,
            'pending_invoices': pending_invoices
        },
        'all_patients': all_patients,
        'all_invoices': all_invoices
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def api_invoice_pdf(request, pk):
    """Download invoice as PDF (public access for now)"""
    try:
        invoice = Invoice.objects.get(id=pk)
    except Invoice.DoesNotExist:
        return Response({'error': 'Invoice not found'}, status=status.HTTP_404_NOT_FOUND)

    if not HAS_REPORTLAB:
        # Fallback: return HTML that browser can print/save as PDF
        html_string = render_to_string('clinic/invoice_pdf.html', {'invoice': invoice})
        return HttpResponse(html_string, content_type='text/html')

    # Generate PDF using ReportLab
    from io import BytesIO

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#198754'),
        spaceAfter=10,
        alignment=1
    )
    elements.append(Paragraph('🦷 INVOICE', title_style))
    elements.append(Paragraph(f'Invoice #{invoice.id}', ParagraphStyle('subtitle', parent=styles['Normal'], fontSize=12, alignment=1)))
    elements.append(Spacer(1, 0.3*inch))

    # Patient Info
    patient_style = ParagraphStyle('patient', parent=styles['Normal'], fontSize=10, leading=12)
    elements.append(Paragraph(f'<b>Patient:</b> {invoice.patient.first_name} {invoice.patient.last_name}', patient_style))
    elements.append(Paragraph(f'<b>Contact:</b> {invoice.patient.contact_number}', patient_style))
    elements.append(Paragraph(f'<b>Date:</b> {invoice.date_created.strftime("%B %d, %Y")}', patient_style))
    elements.append(Spacer(1, 0.3*inch))

    # Services Table
    table_data = [['Service', 'Price', 'Qty', 'Amount']]
    for item in invoice.items.all():
        table_data.append([
            item.service_name_at_time,
            f'₱{float(item.price_at_time):.2f}',
            str(item.quantity),
            f'₱{float(item.total_price()):.2f}'
        ])

    # Add total row
    total_amount = sum(float(item.total_price()) for item in invoice.items.all())
    table_data.append(['', '', 'TOTAL:', f'₱{total_amount:.2f}'])

    table = Table(table_data, colWidths=[2.5*inch, 1*inch, 0.8*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 1), (-1, -2), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -2), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))

    # Footer
    footer_style = ParagraphStyle('footer', parent=styles['Normal'], fontSize=9, textColor=colors.grey, alignment=1)
    elements.append(Paragraph('Thank you for choosing our dental clinic!', footer_style))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Invoice_{invoice.id}_{invoice.patient.first_name}.pdf"'

    return response


@api_view(['GET'])
@authentication_classes([TokenAuthentication, SessionAuthentication])
@permission_classes([permissions.IsAuthenticated])
def api_sales_summary(request):
    """Return sales totals for Day, Week, Month, Year"""
    try:
        from django.utils import timezone
        from datetime import timedelta
        from django.db.models import Sum
        from decimal import Decimal

        now = timezone.now()
        today = now.date()

        # Helper to convert Decimal to float
        def to_float(value):
            if value is None:
                return 0.0
            if isinstance(value, Decimal):
                return float(value)
            return float(value)

        # Today
        today_invoices = Invoice.objects.filter(date_created__date=today)
        today_result = today_invoices.aggregate(Sum('items__price_at_time'))
        today_total = to_float(today_result.get('items__price_at_time__sum'))

        # This week (last 7 days)
        week_start = now - timedelta(days=7)
        week_invoices = Invoice.objects.filter(date_created__gte=week_start)
        week_result = week_invoices.aggregate(Sum('items__price_at_time'))
        week_total = to_float(week_result.get('items__price_at_time__sum'))

        # This month
        month_invoices = Invoice.objects.filter(date_created__year=now.year, date_created__month=now.month)
        month_result = month_invoices.aggregate(Sum('items__price_at_time'))
        month_total = to_float(month_result.get('items__price_at_time__sum'))

        # This year
        year_invoices = Invoice.objects.filter(date_created__year=now.year)
        year_result = year_invoices.aggregate(Sum('items__price_at_time'))
        year_total = to_float(year_result.get('items__price_at_time__sum'))

        # Daily data for chart (last 30 days)
        daily_data = []
        for i in range(29, -1, -1):
            day = now - timedelta(days=i)
            day_date = day.date()
            day_invoices = Invoice.objects.filter(date_created__date=day_date)
            day_result = day_invoices.aggregate(Sum('items__price_at_time'))
            day_total = to_float(day_result.get('items__price_at_time__sum'))
            daily_data.append({
                'date': day_date.isoformat(),
                'total': day_total,
            })

        return Response({
            'today': today_total,
            'week': week_total,
            'month': month_total,
            'year': year_total,
            'daily_chart': daily_data,
        })
    except Exception as e:
        import traceback
        print(f"Error in api_sales_summary: {e}")
        traceback.print_exc()
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
